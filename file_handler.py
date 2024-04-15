import os 
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime

class FileHandler:
    # Handles file operations for the banking system
    def __init__(self):
        # Initialize paths for user data files
        self.users_folder = "./users"
        self.users_file = "./users.txt"
        self.pins_file = "./pins.txt"
        self.balances_file = "./balances.txt" 

    def create_users_folder(self):
        # Create a folder to store individual user folders if it doesn't exist
        if not os.path.exists(self.users_folder):
            os.makedirs(self.users_folder)

    # Users handling
    def create_users_file(self):
        # Create a file to store usernames
        open(self.users_file, 'a').close()

    def add_user_to_file(self, username):
        # Add a new username to the users file
        with open(self.users_file, 'a') as file:
            file.write(username + '\n')

    def get_usernames_from_file(self):
        # Retrieve all usernames from the users file
        with open(self.users_file, 'r') as file:
            return file.read().splitlines()

    # PINs handling  
    def create_pins_file(self):
        # Create a file to store user PINs
        open(self.pins_file, 'a').close()

    
    def add_pin_to_file(self, pin):
        # Add a new PIN to the pins file
        with open(self.pins_file, 'a') as file:
            file.write(pin + '\n')

        
    def get_pins_from_file(self):
        # Retrieve all PINs from the pins file
        with open(self.pins_file, 'r') as file:
            return file.read().splitlines()
    

    # Balances handling
    def create_balances_file(self):
        # Create a file to store user balances
        open(self.balances_file, 'a').close()

    def add_balance_to_file(self, username, balance):
        # Add a user's balance to the balances file
        with open(self.balances_file, 'a') as file:
            file.write(f"{username}:{balance}\n")

    def get_balances_from_file(self):
        # Retrieve all balances from the balances file
        balances = {}
        with open(self.balances_file, 'r') as file:
            for line in file:
                username, balance = line.strip().split(':')
                balances[username] = float(balance)
        return balances

    def update_balance_in_file(self, username, new_balance):
        # Update a user's balance in the balances file
        balances = self.get_balances_from_file()
        balances[username] = new_balance
        with open(self.balances_file, 'w') as file:
            for username, balance in balances.items():
                file.write(f"{username}:{balance}\n")

    # Users' files         
    def create_folder(self, username):
        # Create a personal folder for a user
        folder_path = os.path.join(self.users_folder, username)
        os.makedirs(folder_path, exist_ok= True)
    
    def create_user_file(self, username, pin, initial_deposit):
        # Create a text file with initial user information
        folder_path = os.path.join(self.users_folder, username)
        file_path = os.path.join(folder_path, f"{username}_info.txt")
        with open(file_path, 'w') as file:
            file.write(f"Username: {username}\nPIN: {pin}\nInitial Deposit: {initial_deposit}\n")

    def create_excel_sheet(self, username):
        # Create an Excel sheet to record transaction history
        folder_path = os.path.join(self.users_folder, username)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Date', 'Transaction Type', 'Amount', 'Balance'])
        workbook.save(f"{folder_path}/{username}_transactions.xlsx")

    def update_transaction_history(self, username, transaction_details):
        # Update both text and Excel files with transaction history
        user_folder = os.path.join(self.users_folder, username)
        txt_file_path = os.path.join(user_folder, f"{username}_info.txt")
        excel_file_path = os.path.join(user_folder, f"{username}_transactions.xlsx")

        # Update Text file with transaction details
        with open(txt_file_path, 'a') as file:
            file.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Transaction: {transaction_details[0]}, Amount: {transaction_details[1]}, Balance: {transaction_details[2]}\n")

        # update Excel file with transaction details
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active
        transaction_row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        transaction_row.extend(transaction_details)
        sheet.append(transaction_row)
        workbook.save(excel_file_path)



